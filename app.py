

import os, json, time, sys
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from io import BytesIO
from flask_migrate import Migrate
from flask import send_file
import requests
from dotenv import load_dotenv
from trendyol_api import (
    get_orders, update_package_status, get_order_detail, resolve_line_image,
    get_all_questions, answer_question
)
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User
from models import ShippingLog
from flask import session
from collections import defaultdict
from flask import render_template, request, jsonify
from models import db, Order, OrderItem

# ===========================
#  PAKETLEME MODÜLÜ IMPORT
# ===========================
import io
import json
import pandas as pd
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime, timedelta
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime, timedelta

CACHED_CREATED_ORDERS = []
CACHED_AT = None

def get_created_orders_cached(force_refresh=False, ttl_sec=5):
    """
    Created siparişleri cache'li olarak döner.
    Dashboard ve kargo-toplama aynı kaynağı kullanır.
    """
    global CACHED_CREATED_ORDERS, CACHED_AT

    now = datetime.utcnow()

    cache_expired = (
        CACHED_AT is None or
        (now - CACHED_AT) > timedelta(seconds=ttl_sec)
    )

    if force_refresh or cache_expired or not CACHED_CREATED_ORDERS:
        orders, total = get_orders(status="Created", size=1000)
        CACHED_CREATED_ORDERS = orders
        CACHED_AT = now
        print("🔄 Created siparişler Trendyol'dan yenilendi:", len(CACHED_CREATED_ORDERS))
    else:
        print("⚡ Created siparişler cache'ten alındı:", len(CACHED_CREATED_ORDERS))

    return CACHED_CREATED_ORDERS

try:
    from barcode import Code128
    from barcode.writer import ImageWriter
    HAS_BARCODE = True
except:
    HAS_BARCODE = False

XML_FILE = Path("Entegra.xml")
LOG_FILE = Path("print_log_web.json")

DEFAULT_LAYOUT = {
    "dpi": 203,
    "label_width_mm": 50.0,
    "label_height_mm": 30.0,
    "margin_x_mm": 3.0,
    "margin_top_mm": 3.0,
    "margin_bottom_mm": 3.0,
    "barcode_height_mm": 18.0,
    "module_width": 0.35,
    "header_font_size": 18,
    "barcode_text_font_size": 14,
    "product_font_size": 20,
}
# ===========================
#   XML OKUMA
# ===========================
def read_xml_file():
    if not XML_FILE.exists():
        return pd.DataFrame(columns=["Barkod", "StokKodu", "UrunAdi"])

    try:
        tree = ET.parse(str(XML_FILE))
        root = tree.getroot()
    except Exception as e:
        print("XML okunamadı:", e)
        return pd.DataFrame(columns=["Barkod", "StokKodu", "UrunAdi"])

    rows = []

    for tag in ["product", "Product", "urun", "Urun", "URUN"]:
        for p in root.findall(f".//{tag}"):
            barkod = p.findtext("Barkod") or p.findtext("barcode") or ""
            stok = p.findtext("StokKodu") or p.findtext("Kod") or ""
            urun = p.findtext("UrunAdi") or p.findtext("Baslik") or ""
            rows.append({"Barkod": barkod, "StokKodu": stok, "UrunAdi": urun})

    return pd.DataFrame(rows)
# ---- Flask App ----
import os
from dotenv import load_dotenv

load_dotenv()

import sys, os
from flask import Flask

# PyInstaller uyumlu base path
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS  # derlenmiş exe içindeki temp klasör
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---- Flask Ayarları ----
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, "templates"),
            static_folder=os.path.join(BASE_DIR, "static"))
app.secret_key = os.getenv("SECRET_KEY", "supersecret")
# ⬇⬇⬇ BUNLAR BURAYA GELECEK ⬇⬇⬇
from yakamel_paketleme import paketleme_blueprint
app.register_blueprint(paketleme_blueprint)
# ⬆⬆⬆ BUNLAR BURAYA GELECEK ⬆⬆⬆
# ---- DB Ayarları ----
DATABASE_URL = os.getenv("DATABASE_URL")

app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://postgres:12345@localhost:5432/trendyol_v2_2025_program"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)
migrate = Migrate(app, db)


# ---- Login Manager ----
login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# ---- .env yükle ----
load_dotenv()

PAGE_SIZE = 20

# ===========================
#   LOG SİSTEMİ
# ===========================
def ensure_log():
    if not LOG_FILE.exists():
        LOG_FILE.write_text("[]", encoding="utf-8")


def add_log(barcode, qty, stok, urun):
    ensure_log()
    try:
        data = json.loads(LOG_FILE.read_text("utf-8"))
    except:
        data = []

    for _ in range(int(qty)):   # 🔥 HER ETİKET AYRI
        data.append({
            "ts": datetime.now().isoformat(),  # TR saati
            "barcode": barcode,
            "qty": 1,
            "stok": stok,
            "urun": urun
        })

    LOG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")

def get_today_count():
    ensure_log()
    try:
        data = json.loads(LOG_FILE.read_text("utf-8"))
    except:
        return 0

    today = datetime.now().date()   # Yerel gün
    total = 0

    for r in data:
        try:
            ts = datetime.fromisoformat(r["ts"])
            if ts.date() == today:
                total += 1           # Her log = 1 etiket
        except:
            pass

    return total

# ===========================
#  ETİKET OLUŞTURMA
# ===========================
def mm_to_px(mm, dpi):
    return int((mm / 25.4) * dpi)

def load_font(size, bold=False):
    paths = [
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold
        else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for p in paths:
        if Path(p).exists():
            return ImageFont.truetype(p, size)
    return ImageFont.load_default()

def render_label(barcode, stok, urun):
    dpi = DEFAULT_LAYOUT["dpi"]
    W = mm_to_px(DEFAULT_LAYOUT["label_width_mm"], dpi)
    H = mm_to_px(DEFAULT_LAYOUT["label_height_mm"], dpi)

    img = Image.new("L", (W, H), 255)
    d = ImageDraw.Draw(img)

    # Stok kodu
    font_h = load_font(DEFAULT_LAYOUT["header_font_size"], bold=True)
    d.text((10, 5), f"STOK: {stok}", 0, font=font_h)

    # Barkod
    if HAS_BARCODE:
        code = Code128(str(barcode), writer=ImageWriter())
        tmp = io.BytesIO()
        code.write(tmp, options={"module_width": DEFAULT_LAYOUT["module_width"], "font_size": 0})
        bc = Image.open(io.BytesIO(tmp.getvalue()))

        bc = bc.resize((W - 20, 50))
        img.paste(bc, (10, 40))

    # Ürün
    font_u = load_font(DEFAULT_LAYOUT["product_font_size"], bold=True)
    d.text((10, H - 35), urun[:25], 0, font=font_u)

    return img
# ===========================
#  PAKETLEME SAYFASI
# ===========================
@app.route("/paketleme", methods=["GET"])
@login_required
def paketleme():
    ensure_log()

    xml_name = XML_FILE.name if XML_FILE.exists() else None
    q = request.args.get("q", "").strip().lower()
    page = int(request.args.get("page", 1))
    per_page = 20
    counter = get_today_count()

    # Eğer arama yoksa boş liste göster
    if q == "":
        return render_template(
            "paketleme.html",
            xml_name=xml_name,
            rows=[],
            q=q,
            counter=counter,
            total_pages=0,
            page=1
        )

    # Tüm XML'i oku
    df = read_xml_file()

    # lowercase kolonlar
    df["lb"] = df["Barkod"].astype(str).str.lower()
    df["ls"] = df["StokKodu"].astype(str).str.lower()
    df["lu"] = df["UrunAdi"].astype(str).str.lower()

    # Filtre
    filt = df[
        df["lb"].str.contains(q) |
        df["ls"].str.contains(q) |
        df["lu"].str.contains(q)
    ]

    total_rows = len(filt)
    total_pages = (total_rows // per_page) + (1 if total_rows % per_page else 0)

    # Sayfa sınırları
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages

    start = (page - 1) * per_page
    end = start + per_page

    filt_page = filt.iloc[start:end]

    rows = filt_page[["Barkod", "StokKodu", "UrunAdi"]].to_dict(orient="records")

    return render_template(
        "paketleme.html",
        xml_name=xml_name,
        rows=rows,
        q=q,
        counter=counter,
        page=page,
        total_pages=total_pages
    )


@app.route("/upload_xml", methods=["POST"])
@login_required
def upload_xml():
    f = request.files.get("xml_file")
    if not f:
        flash("XML seçilmedi!", "err")
        return redirect("/paketleme")

    f.save(XML_FILE)
    flash("XML başarıyla yüklendi!", "ok")
    return redirect("/paketleme")


@app.route("/preview")
@login_required
def preview_label_route():
    barcode = request.args.get("barcode")
    stok = request.args.get("stok_kodu")
    urun = request.args.get("urun_adi")

    img = render_label(barcode, stok, urun)
    buf = io.BytesIO()
    img = img.resize((img.width*3, img.height*3))
    img.save(buf, "PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


@app.route("/browser_print", methods=["POST"])
@login_required
def browser_print():
    barcode = request.form.get("barcode")
    stok = request.form.get("stok_kodu")
    urun = request.form.get("urun_adi")
    qty = int(request.form.get("qty", "1"))

    add_log(barcode, qty, stok, urun)

    return render_template_string("""
        <html><body onload="window.print();window.close();">
        {% for i in range(qty) %}
            <img src="/preview?barcode={{barcode}}&stok_kodu={{stok}}&urun_adi={{urun}}">
        {% endfor %}
        </body></html>
    """, barcode=barcode, stok=stok, urun=urun, qty=qty)

SURAT_KARGO_HESAPLARI = {
    "564724": {  # RUNADES
        "KullaniciAdi": "1500205406",   # ✅ sözleşme kodu artık kullanıcı adı
        "Sifre": "Yunus.5406",          # ✅ senin gerçek şifren
        "SozlesmeKodu": "1500205406",   # aynı kalabilir
        "FirmaAdi": "YUNUS EMRE KAYA"
    },
    "940685": {  # YAKAMEL TEKSTİL - TUĞÇE YILMAZ
        "KullaniciAdi": "1500204598",
        "Sifre": "Yunus.5406",
        "SozlesmeKodu": "1500204598",
        "FirmaAdi": "TUĞÇE YILMAZ"
    },
    "1086036": {  # CMZ COLLECTION
        "KullaniciAdi": "1500200828",
        "Sifre": "Yunus.5406",
        "SozlesmeKodu": "1500200828",
        "FirmaAdi": "CMZ COLLECTION TEKSTİL"
    },
    "1127426": {  # BARLİZ TEKSTİL
        "KullaniciAdi": "1500199645",
        "Sifre": "Yunus.5406",
        "SozlesmeKodu": "1500199645",
        "FirmaAdi": "BARLİZ TEKSTİL"
    },
    "938355": {  # YKML-YAŞAR YILMAZ
        "KullaniciAdi": "1500229286",
        "Sifre": "Yunus.5406",
        "SozlesmeKodu": "1500229286",
        "FirmaAdi": "YKML - YAŞAR YILMAZ"
    },
    "994330": {  # BAY BAYAN
        "KullaniciAdi": "1500228013",
        "Sifre": "Yunus.5406",
        "SozlesmeKodu": "1500228013",
        "FirmaAdi": "BAY BAYAN TEKSTİL"
    }
}

# 🔹 Filtre SKU listesi
FILTER_SKUS = [
    "KFTK", "ETK3I", "BSKLE", "KIKT", "ETKP", "TAYT", "ESF3I", "ESPE", "SWT3I", "PLZO",
    "KSKP", "ESFKP", "KMTK", "BKTK", "KKTK", "OFBS", "BTSH", "SBP", "SGP", "UBP", "UGP",
    "KBP", "KGP", "ULP", "KKFE", "BSKLTY", "TSH", "HRKA", "FDKY", "FSAH", "KSTK", "OFTA",
    "HRTK", "EPA", "OBSWT", "DYTK", "SLP", "KLP", "ELBS", "DKP", "KMNO", "ESTK", "SAL",
    "BAT", "HRKI", "CNT", "MTR", "PBK", "OFT", "PLR"
]
FILTER_SKUS = [sku.upper() for sku in FILTER_SKUS]

from datetime import datetime, timezone

# ⬇️ BURANIN ALTINA EKLE ⬇️

# 🔹 Mağaza ve Renk Filtresi Ayarları
AVAILABLE_SUPPLIERS = {
    "564724": "RUNADES",
    "940685": "YAKAMEL TEKSTİL",
    "938355": "YKML",
    "1086036": "CMZ COLLECTION",
    "1127426": "BARLİZ TEKSTİL",
    "994330": "BAY BAYAN"
}

COLOR_FILTERS = ["SİYAH", "LACİVERT", "FÜME", "KAHVERENGİ", "HAKİ", "BEYAZ", "BEJ", "GRİ", "KIRMIZI", "MAVİ", "YEŞİL"]

def parse_date(dt):
    """Trendyol tarih alanlarını güvenli şekilde datetime objesine çevirir (UTC aware)"""
    if not dt:
        return None
    try:
        if isinstance(dt, str) and dt.isdigit():
            dt = int(dt)

        if isinstance(dt, (int, float)):
            # Trendyol timestamp milisaniye cinsinden geliyor → UTC
            return datetime.fromtimestamp(dt / 1000.0, tz=timezone.utc)
        elif isinstance(dt, str):
            # ISO string format (ör: "2025-10-01T08:55:42.000Z")
            return datetime.fromisoformat(dt.replace("Z", "+00:00")).astimezone(timezone.utc)
        elif isinstance(dt, datetime):
            # Eğer timezone bilgisi yoksa UTC ata
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception as e:
        print("⚠️ Tarih parse edilemedi:", dt, e)
    return None

# 🔹 Siparişleri filtreleyen fonksiyon
def filter_orders(orders):
    filtered = []
    for order in orders:
        new_lines = []
        for line in order.get("lines", []):
            sku = (line.get("merchantSku") or line.get("sku") or "").upper()
            if sku in FILTER_SKUS:
                new_lines.append(line)
        if new_lines:
            order["lines"] = new_lines
            filtered.append(order)
    return filtered
# 🔹 Trendyol kargo bildirimi fonksiyonu (BURAYA EKLE)
def bildir_trendyol_kargo(supplier_id, package_id, tracking_number):
    url = f"https://api.trendyol.com/sapigw/suppliers/{supplier_id}/shipment-packages"
    payload = [{
        "id": package_id,
        "trackingNumber": tracking_number,
        "shipmentProviderId": 3,  # 3 = Sürat Kargo
        "status": "Shipped"
    }]
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Basic <API_KEY:SECRET base64>"  # bunu .env'den de çekebiliriz
    }
    r = requests.put(url, json=payload, headers=headers, timeout=15)
    print("📨 Trendyol Kargo Bildirimi:", r.status_code, r.text)
    return r.status_code == 200


# ---- Ana Menü ----
from datetime import datetime, timezone, timedelta
IST = timezone(timedelta(hours=3))  # Türkiye saati

@app.route("/")
def index():
    try:
        today = datetime.now(IST).date()

        created_orders, _ = get_orders(status="Created", size=200)
        created_count = len(created_orders)

        picking_orders, _ = get_orders(status="Picking", size=200)
        picking_count = len(picking_orders)

        shipped_orders, _ = get_orders(status="Shipped", size=200)
        shipped_count = len(shipped_orders)

        # Yapılması gereken kargo
        to_ship = created_count
        # 🔹 Bugün taşımada olanları yakala
        daily_shipped = []

        for o in shipped_orders:

            dt_parsed = parse_date(o.get("shipmentCreatedDate"))

            if not dt_parsed:
                dt_parsed = parse_date(o.get("lastModifiedDate") or o.get("orderDate"))

            if dt_parsed:

                if dt_parsed.tzinfo is None:
                    dt_parsed = dt_parsed.replace(tzinfo=timezone.utc)

                dt_local = dt_parsed.astimezone(IST)

                if dt_local.date() == today:
                    daily_shipped.append(o)

        # 🔹 bugün gönderilen
        shipped_today_count = len(daily_shipped)

        # 📦 GERÇEK TOPLAM
        total_all = created_count + picking_count + shipped_count

    except Exception as e:

        print("❌ Kargo istatistikleri alınamadı:", e)

        created_count = 0
        picking_count = 0
        shipped_count = 0
        shipped_today_count = 0
        total_all = 0

    return render_template(
        "index.html",
        created_count=to_ship,
        picking_count=picking_count,
        shipped_count=shipped_today_count,
        total_all=to_ship + shipped_today_count
    )
# ============================
# 🚀 D A S H B O A R D – MODEL A (SABİT SAYFALAMA)
# ============================
@app.route("/dashboard")
@login_required
def dashboard():

    status = request.args.get("status", "Created")
    page = int(request.args.get("page", 1))
    per_page = 100

    supplier_filter = request.args.get("supplier", "")
    color_filter = request.args.get("color", "")
    selected_filters = request.args.getlist("filter")

    # 🔥 KRİTİK FIX
    selected_filters = [f for f in selected_filters if f and f != "ALL"]

    # 🔹 Siparişleri çek
    all_orders = get_created_orders_cached()
    real_total_to_ship = len(all_orders)

    orders_raw = []

    for o in all_orders:

        if not isinstance(o, dict):
            continue

        # ✅ ID FIX
        o["id"] = (
            o.get("shipmentPackageId")
            or o.get("packageId")
            or o.get("id")
        )

        # ✅ SUPPLIER FIX
        o["supplier_id"] = o.get("supplier_id") or o.get("supplierId")

        if o["id"]:
            orders_raw.append(o)

    # -----------------------------
    # MAĞAZA FİLTRESİ
    # -----------------------------
    if supplier_filter and supplier_filter != "ALL":
        orders_raw = [
            o for o in orders_raw
            if str(o.get("supplier_id", "")) == supplier_filter
        ]

    # -----------------------------
    # RENK FİLTRESİ (FIX)
    # -----------------------------
    if color_filter and color_filter != "":
        filtered = []
        for o in orders_raw:
            for l in o.get("lines", []):
                if color_filter.upper() in str(l.get("productColor","")).upper():
                    filtered.append(o)
                    break
        orders_raw = filtered

    # -----------------------------
    # SKU FİLTRESİ (FIX)
    # -----------------------------
    selected_filters = request.args.getlist("filter")

    # 🔥 GERÇEK FIX
    if selected_filters == [''] or selected_filters == []:
        selected_filters = []
    if selected_filters and len(selected_filters) > 0:
        filtered = []
        for o in orders_raw:
            for l in o.get("lines", []):
                sku = (l.get("merchantSku") or l.get("sku") or "").upper()

                if any(f in sku for f in selected_filters):
                    filtered.append(o)
                    break
        orders_raw = filtered

    # -----------------------------
    # SIRALAMA
    # -----------------------------
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)

    def deadline_sort(order):
        dl = order.get("extendedAgreedDeliveryDate") or order.get("agreedDeliveryDate")

        if not dl:
            return 999999999

        try:
            dt = parse_date(dl)
            if not dt:
                return 999999999

            return (dt - now).total_seconds()
        except:
            return 999999999

    orders_raw.sort(key=deadline_sort)

    # -----------------------------
    # 24 SAAT
    # -----------------------------
    urgent_24h = 0

    for o in orders_raw:
        dl = o.get("extendedAgreedDeliveryDate") or o.get("agreedDeliveryDate")

        if not dl:
            continue

        try:
            dt_deadline = parse_date(dl)
            if not dt_deadline:
                continue

            kalan = (dt_deadline - now).total_seconds()

            if 0 < kalan <= 86400:
                urgent_24h += 1
        except:
            pass

    # -----------------------------
    # MAĞAZA ADI
    # -----------------------------
    for o in orders_raw:
        o["supplier_name"] = AVAILABLE_SUPPLIERS.get(
            str(o.get("supplier_id")), "Bilinmeyen"
        )

    # -----------------------------
    # SAYFALAMA
    # -----------------------------
    total_orders = len(orders_raw)

    total_pages = max((total_orders // per_page) + (1 if total_orders % per_page else 0), 1)

    page = max(1, min(page, total_pages))

    start = (page - 1) * per_page
    end = start + per_page

    orders = orders_raw[start:end]

    start_page = max(1, page - 3)
    end_page = min(total_pages, page + 3)

    page_numbers = list(range(start_page, end_page + 1))

    return render_template(
        "dashboard.html",
        orders=orders,
        page=page,
        total_pages=total_pages,
        page_numbers=page_numbers,
        total_to_ship=real_total_to_ship,
        urgent_24h=urgent_24h,
        selected_filters=selected_filters,
        supplier_filter=supplier_filter,
        color_filter=color_filter,
        status=status,
        per_page=per_page,
        current_filters={
            "supplier": supplier_filter,
            "color": color_filter,
            "filter": selected_filters,
            "status": status
        }
    )
# --------------------------------
# TOPLAMA SAYFASI
# --------------------------------
from trendyol_api import resolve_line_image

SELLER_NAMES = {
    "994330": "BAY BAYAN",
    "938355": "YKML",
    "1127426": "BARLİZ",
    "1086036": "CMZ COLLECTION",
    "940685": "YAKAMEL TEKSTİL",
    "564724": "RUNADES"
}

@app.route("/toplama")
def toplama():

    orders = get_created_orders_cached()

    waiting_orders = []
    picked_orders = []

    for o in orders:

        package_id = str(
            o.get("shipmentPackageId")
            or o.get("packageId")
            or o.get("id")
        )
        o["packageId"] = package_id

        o["customerFirstName"] = o.get("customerFirstName", "")
        o["customerLastName"] = o.get("customerLastName", "")

        supplier_id = str(
            o.get("supplier_id")
            or o.get("supplierId")
            or o.get("sellerId")
            or ""
        )

        o["supplierName"] = SELLER_NAMES.get(supplier_id, supplier_id)

        lines = o.get("lines", [])

        for line in lines:

            barcode = str(line.get("barcode") or "")

            # 1️⃣ sipariş içinden resim varsa
            image = (
                line.get("imageUrl")
                or line.get("productImage")
                or line.get("productImageUrl")
            )

            # 2️⃣ cache kontrol
            if not image and barcode in IMAGE_CACHE:
                image = IMAGE_CACHE[barcode]

            # 3️⃣ product API (sadece ilk sefer)
            if not image:

                try:
                    image = resolve_line_image(
                        supplier_id=supplier_id,
                        barcode=barcode,
                        merchantSku=line.get("merchantSku") or line.get("stockCode"),
                        sku=line.get("sku"),
                        productCode=str(line.get("productCode") or "")
                    )

                    if image:
                        IMAGE_CACHE[barcode] = image

                except:
                    image = None

            # 4️⃣ placeholder
            if not image:
                image = "https://via.placeholder.com/90"

            line["productImage"] = image

            line["productColor"] = (
                line.get("productColor")
                or line.get("color")
                or "-"
            )

            line["productSize"] = (
                line.get("productSize")
                or line.get("size")
                or "-"
            )

            line["productName"] = line.get("productName", "")

            line["quantity"] = line.get("quantity", 1)

        o["lines"] = lines

        existing = Order.query.filter_by(package_id=package_id).first()

        if existing and existing.order_stage == "picked":
            picked_orders.append(o)
        else:
            waiting_orders.append(o)

    return render_template(
        "toplama.html",
        waiting_orders=waiting_orders,
        picked_orders=picked_orders
    )
# --------------------------------
# BARKOD OKUTMA API
# --------------------------------
@app.route("/scan_barcode", methods=["POST"])
def scan_barcode():

    data = request.json
    order_id = data["order_id"]
    barcode = data["barcode"]

    item = (
        OrderItem.query
        .filter_by(order_id=order_id, barcode=barcode)
        .filter(OrderItem.picked_qty < OrderItem.quantity)
        .first()
    )

    if not item:
        return jsonify({"error": "Bu barkod siparişte yok"})

    item.picked_qty += 1
    db.session.commit()

    return jsonify({
        "success": True,
        "product": item.product_name,
        "picked": item.picked_qty,
        "total": item.quantity
    })

# --- IMAGE CACHE ---
IMAGE_CACHE = {}

def preload_images():
    """
    Sunucu açılırken siparişlerdeki ürün resimlerini cache'e alır.
    Aynı barkod tekrar API çağırmaz.
    """
    try:
        orders = get_created_orders_cached()
        for o in orders:
            supplier_id = str(o.get("supplier_id") or o.get("supplierId") or o.get("sellerId") or "")
            for line in o.get("lines", []):
                barcode = line.get("barcode")
                if not barcode:
                    continue
                if barcode in IMAGE_CACHE:
                    continue

                img = resolve_line_image(
                    supplier_id=supplier_id,
                    barcode=barcode,
                    merchantSku=line.get("merchantSku"),
                    sku=line.get("sku"),
                    productCode=str(line.get("productCode") or "")
                )
                if img:
                    IMAGE_CACHE[barcode] = img
    except Exception as e:
        print("❌ preload_images error:", e)
# ---- Sorular ----
@app.route("/questions")
@login_required
def questions():
    try:
        # Trendyol API’den ürün ve sipariş sorularını çek
        product_questions, order_questions = get_all_questions(
            status="WAITING_FOR_ANSWER", days=14
        )
        return render_template(
            "questions.html",
            product_questions=product_questions,
            order_questions=order_questions
        )
    except Exception as e:
        flash(f"Sorular alınamadı: {e}", "danger")
        return redirect(url_for("index"))


# ---- Kullanıcı Kayıt ----
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if User.query.filter_by(username=username).first():
            flash("❌ Bu kullanıcı adı zaten var.", "danger")
            return redirect(url_for("register"))

        new_user = User(username=username)
        new_user.set_password(password)   # ✔️ HASHLİYOR

        db.session.add(new_user)
        db.session.commit()

        flash("✔️ Kayıt başarılı, giriş yapabilirsiniz", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


# ---- Kullanıcı Giriş ----
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for("index"))
        else:
            flash("Hatalı kullanıcı adı veya şifre", "danger")

    return render_template("login.html")


# ---- Çıkış ----
@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))
# ---- Şifre Değiştir ----
@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        old_password = request.form.get("old_password")
        new_password = request.form.get("new_password")

        if not check_password_hash(current_user.password, old_password):
            flash("❌ Mevcut şifre yanlış!", "danger")
            return redirect(url_for("change_password"))

        current_user.set_password(new_password)
        db.session.commit()
        flash("✅ Şifreniz başarıyla güncellendi.", "success")
        return redirect(url_for("index"))

    return render_template("change_password.html")

# ---- Soruya Cevap Ver ----
@app.route("/cevapla/<question_id>", methods=["POST"])
@login_required
def cevapla(question_id):
    if current_user.role not in ["soru", "ofis", "admin"]:
        flash("❌ Bu soruya cevap verme yetkiniz yok.", "danger")
        return redirect(url_for("questions"))

    cevap_text = request.form.get("cevap")
    supplier_id = request.form.get("supplier_id")

    if not cevap_text or len(cevap_text) < 10:
        flash("⚠️ Cevap en az 10 karakter olmalı.", "warning")
        return redirect(url_for("questions"))

    ok = answer_question(supplier_id, question_id, cevap_text)

    if ok:
        flash("✅ Cevabınız başarıyla gönderildi.", "success")
    else:
        flash("❌ Cevap gönderilemedi.", "danger")

    return redirect(url_for("questions"))

# ---- Cevaplanan Sorular ----
@app.route("/cevaplanan-sorular")
@login_required
def cevaplanan_sorular():
    if current_user.role not in ["soru", "ofis", "admin"]:
        flash("❌ Bu sayfayı görüntüleme yetkiniz yok.", "danger")
        return redirect(url_for("index"))

    product_questions, order_questions = get_all_questions(status="ANSWERED", days=14)
    sorular = [s for s in product_questions + order_questions if s.get("answerText")]
    return render_template("cevaplanan_sorular.html", sorular=sorular)

# ---- Admin Panel ----
@app.route("/admin_panel")
@login_required
def admin_panel():
    if current_user.role != "admin":
        flash("❌ Admin paneline giriş yetkiniz yok.", "danger")
        return redirect(url_for("index"))

    users = User.query.all()
    return render_template("admin_panel.html", users=users)

# ---- Rol Değiştirme ----
@app.route("/change_role/<int:user_id>", methods=["POST"])
@login_required
def change_role(user_id):
    if current_user.role != "admin":
        flash("❌ Rol değiştirme yetkiniz yok.", "danger")
        return redirect(url_for("index"))

    user = User.query.get(user_id)
    if not user:
        flash("❌ Kullanıcı bulunamadı.", "danger")
        return redirect(url_for("admin_panel"))

    new_role = request.form.get("role")
    if new_role not in ["üye", "kargo", "soru", "ofis", "admin"]:
        flash("❌ Geçersiz rol seçildi.", "danger")
        return redirect(url_for("admin_panel"))

    user.role = new_role
    db.session.commit()
    flash(f"✅ {user.username} kullanıcısının rolü '{new_role}' olarak güncellendi.", "success")
    return redirect(url_for("admin_panel"))
# ---- Şifre Sıfırlama (Admin) ----
@app.route("/reset_password/<int:user_id>", methods=["POST"])
@login_required
def reset_password(user_id):
    if current_user.role != "admin":
        flash("❌ Şifre sıfırlama yetkiniz yok.", "danger")
        return redirect(url_for("admin_panel"))

    user = User.query.get(user_id)
    if not user:
        flash("❌ Kullanıcı bulunamadı.", "danger")
        return redirect(url_for("admin_panel"))

    new_password = request.form.get("new_password")
    if not new_password or len(new_password) < 4:
        flash("⚠️ Yeni şifre en az 4 karakter olmalı.", "warning")
        return redirect(url_for("admin_panel"))

    user.password = generate_password_hash(new_password, method="pbkdf2:sha256")
    db.session.commit()
    flash(f"✅ {user.username} için yeni şifre başarıyla güncellendi.", "success")
    return redirect(url_for("admin_panel"))

# ---- API Endpoints ----
@app.route("/api/orders")
def api_orders():
    status = request.args.get("status", "Created")
    page = int(request.args.get("page", 0))
    size = int(request.args.get("size", PAGE_SIZE))
    orders, total = get_orders(status=status, size=size)
    return jsonify({"orders": orders, "size": size, "total": total})

@app.route("/api/line-image")
def api_line_image():
    supplier_id = request.args.get("supplier_id")
    barcode = request.args.get("barcode")
    merchantSku = request.args.get("merchantSku")
    sku = request.args.get("sku")
    productCode = request.args.get("productCode")

    print("🔍 /api/line-image İSTEĞİ:")
    print("supplier_id:", supplier_id)
    print("barcode:", barcode)
    print("merchantSku:", merchantSku)
    print("sku:", sku)
    print("productCode:", productCode)

    # 🔥 Barkod veya SKU ile resim çöz
    url = resolve_line_image(
        supplier_id=supplier_id,
        barcode=barcode,
        merchantSku=merchantSku,
        sku=sku,
        productCode=productCode
    )

    # ❗ Eğer bulunamazsa default değer dön
    if not url:
        url = "https://via.placeholder.com/300x300.png?text=BARKOD+YOK"

    return jsonify({"url": url})

def get_all_created_orders():
    page = 0
    size = 200
    all_orders = []

    while True:
        orders, total = get_orders(status="Created", page=page, size=size)
        if not orders:
            break

        all_orders.extend(orders)

        if len(orders) < size:
            break

        page += 1

    return all_orders

# ---- Sipariş İşleme ----
@app.route("/isleme-al/<supplier_id>/<int:package_id>", methods=["POST"])
@login_required
def isleme_al(supplier_id, package_id):
    from datetime import datetime, timezone, timedelta
    IST = timezone(timedelta(hours=3))
    if current_user.role not in ["kargo", "ofis", "admin"]:
        flash("❌ Sipariş işleme alma yetkiniz yok.", "danger")
        return redirect(url_for("dashboard"))

    lines_raw = request.form.get("lines", "[]")
    try:
        lines = json.loads(lines_raw)
    except:
        lines = []

    ok = update_package_status(
        supplier_id,
        package_id,
        lines,
        status="Picking"
    )

    print("=== LOG BAŞLIYOR ===")

    if ok:
        print("update_package_status OK ✓")

        # 🔥 Created sipariş cache temizle
        global CACHED_CREATED_ORDERS, CACHED_AT
        CACHED_CREATED_ORDERS = []
        CACHED_AT = None

        from models import ShippingLog
        from trendyol_api import get_order_detail
        from excel_kargo_kaydet import excele_ekle
        from datetime import timezone, timedelta

        try:
            order_detail = get_order_detail(supplier_id, package_id)
            print("order_detail:", order_detail)

            # --- Müşteri ---
            customer_name = f"{order_detail.get('customerFirstName','')} {order_detail.get('customerLastName','')}"

            # --- Mağaza ---
            supplier_name = (
                order_detail.get("supplier_name")
                or order_detail.get("supplierName")
                or ""
            )

            # --- Sipariş ---
            order_number = order_detail.get("orderNumber")
            tracking_number = order_detail.get("cargoTrackingNumber")

            # --- Sipariş Tarihi ---
            IST = timezone(timedelta(hours=3))
            order_date = None
            order_date_raw = order_detail.get("orderDate")

            if order_date_raw:
                try:
                    dt = datetime.fromtimestamp(
                        int(order_date_raw) / 1000,
                        tz=timezone.utc
                    )
                    order_date = dt.astimezone(IST)
                except:
                    order_date = None

            print("→ Line sayısı:", len(lines))

            # =========================
            # DB LOG KAYDI (TEK PAKET = TEK LOG)
            # =========================

            urun_adlari = []
            skular = []
            renkler = []
            bedenler = []
            toplam_adet = 0

            for line in lines:
                urun_adlari.append(str(line.get("productName")))
                skular.append(str(line.get("merchantSku")))
                renkler.append(str(line.get("productColor")))
                bedenler.append(str(line.get("productSize")))
                toplam_adet += int(line.get("quantity", 1))

            log = ShippingLog(
                supplier_id=supplier_id,
                supplier_name=supplier_name,

                order_number=order_number,
                tracking_number=tracking_number,
                package_id=str(package_id),

                customer_name=customer_name,
                order_date=order_date,

                product_name=" | ".join(urun_adlari),
                sku=" | ".join(skular),
                quantity=toplam_adet,
                color=" | ".join(renkler),
                size=" | ".join(bedenler),

                processed_at=datetime.now(IST),
                shipped_at=None
            )

            db.session.add(log)
            db.session.commit()

            # =========================
            # EXCEL KAYDI (TEK SEFER)
            # =========================
            try:
                urunler = []

                for line in lines:
                    urunler.append({
                        "stok_kodu": line.get("merchantSku"),
                        "urun_adi": line.get("productName"),
                        "renk": line.get("productColor"),
                        "beden": line.get("productSize"),
                        "adet": line.get("quantity", 1)
                    })

                excele_ekle(
                    order_no=order_number,
                    kullanici=current_user.username,
                    urunler=urunler
                )

                print("✓ EXCEL KAYDI ALINDI")

            except Exception as excel_err:
                print("❌ EXCEL HATASI:", excel_err)

        except Exception as e:
            print("❌ LOG HATASI:", e)
            db.session.rollback()

    else:
        print("❌ update_package_status başarısız!")

    # =========================
    # PARAMETRELERİ GERİ GÖNDER
    # =========================
    params = {}

    params["page"] = request.form.get("page", "1")
    params["status"] = request.form.get("status", "Created")

    if request.form.get("supplier"):
        params["supplier"] = request.form.get("supplier")

    if request.form.get("color"):
        params["color"] = request.form.get("color")

    for f in request.form.getlist("filter"):
        params.setdefault("filter", []).append(f)

    if request.form.get("urgent"):
        params["urgent"] = request.form.get("urgent")

    if request.form.get("row_index"):
        params["row_index"] = request.form.get("row_index")

    flash("✅ Sipariş işleme alındı", "success")
    return redirect(url_for("dashboard", **params))

# ---- Etiket Yazdır ----
@app.route("/etiket-yazdir/<supplier_id>/<int:package_id>")
@login_required
def etiket_yazdir(supplier_id, package_id):
    order = get_order_detail(supplier_id, package_id)
    if not order:
        flash("Paket detayı getirilemedi.", "danger")
        return redirect(url_for("dashboard"))
    return render_template("etiket.html", o=order)

from collections import defaultdict
import re, unicodedata
from flask import render_template, redirect, url_for, flash
from flask_login import login_required, current_user

@app.route("/kargo-toplama")
@login_required
def kargo_toplama():
    if current_user.role not in ["kargo", "ofis", "admin"]:
        flash("❌ Bu sayfaya erişim yetkiniz yok.", "danger")
        return redirect(url_for("dashboard"))

    try:
        all_orders = get_created_orders_cached(force_refresh=False)
        print("🚛 Kargo toplama created sipariş:", len(all_orders))
        print("API’den gelen created paket:", len(all_orders))

        # 🔹 STK -> Ürün isimleri
        STK_TO_NAME = {
            "ETK3I": "JAGGER EŞOFMAN TAKIMI",
            "BMBRTK": "BOMBER TAKIM",
            "HRKA": "HIRKA",
            "FDKY": "DİK YAKA",
            "KFTK": "KADIN FİTİLLİ TAKIM",
            "BSKLE": "BİSİKLET YAKA TAKIM",
            "SWT3I": "SWEATSHİRT",
            "ESF3I": "3 İPLİK TEK ALT",
            "ESPE": "PENYE EŞOFMAN ALTI",
            "KMTK": "KİMONO BÜRÜMCÜK TAKIM",
            "BKTK": "BÜRÜMCÜK KISA KOLLU TAKIM",
            "KKTK": "KAŞKORSE TAKIM",
            "DYTK": "DİK YAKA EŞOFMAN TAKIMI",
            "PLR": "POLAR HIRKA"
        }

        # 🔹 Renk normalize fonksiyonu
        def normalize_color_name(name):
            if not name:
                return {"kod": "#cccccc", "ad": "Belirsiz", "key": "belirsiz"}

            raw = unicodedata.normalize("NFKD", name)
            raw = raw.replace("İ", "i").replace("ı", "i").replace("i̇", "i")
            raw = raw.encode("ascii", "ignore").decode("utf-8", "ignore")
            raw = raw.strip().upper()

            raw = re.sub(r"\((.*?)\)", "", raw)
            raw = re.sub(r"\s+", " ", raw)

            for junk in ["RENK", "RENGI", "RENGİ", "RNG", "MAVI", "MAVİ", "MAVISI", "MAVİSİ", "VERT", "MELANJ", "COLOR"]:
                raw = raw.replace(junk, "")

            if re.search(r"FÜM|FUME|SMOKE|CHARCOAL|DARK GREY", raw): raw = "FÜME"
            elif re.search(r"GRI|GREY|GRAY", raw): raw = "GRİ"
            elif re.search(r"SAX|SAKS|SAX BLUE|SAKS MAVI", raw): raw = "SAKS MAVİSİ"
            elif re.search(r"BEBE|BABY BLUE|BEBEMAVI", raw): raw = "BEBE MAVİSİ"
            elif re.search(r"MURDUM|MORDO", raw): raw = "MÜRDÜM"
            elif re.search(r"BLACK|SIYAH", raw): raw = "SİYAH"
            elif re.search(r"LACI|LACIVERT", raw): raw = "LACİVERT"
            elif re.search(r"HAKI|KHAKI", raw): raw = "HAKİ"
            elif re.search(r"KAHVE|BROWN", raw): raw = "KAHVERENGİ"
            elif re.search(r"BEIGE", raw): raw = "BEJ"
            elif re.search(r"VIZON|VISON", raw): raw = "VİZON"
            elif re.search(r"BORDO", raw): raw = "BORDO"
            elif re.search(r"ORANGE", raw): raw = "TURUNCU"

            raw = raw.strip()
            renk_ad = raw.title()

            renk_key = (
                raw.lower()
                .replace(" ", "")
                .replace("-", "")
                .replace("_", "")
                .replace(".", "")
                .replace("/", "")
                .replace("\\", "")
                .replace("ı", "i")
                .replace("ş", "s")
                .replace("ç", "c")
                .replace("ö", "o")
                .replace("ü", "u")
                .replace("ğ", "g")
            )

            renkler = {
                "beyaz": "#ffffff",
                "siyah": "#000000",
                "lacivert": "#001f3f",
                "mavi": "#007bff",
                "saksmavisi": "#0066cc",
                "bebemavisi": "#a5d8ff",
                "gri": "#b0b0b0",
                "füme": "#5a5a5a",
                "kirmizi": "#d62828",
                "bordo": "#800020",
                "yesil": "#198754",
                "pembe": "#f472b6",
                "fusya": "#c026d3",
                "mor": "#6d28d9",
                "mürdüm": "#5f0f40",
                "kahverengi": "#6f4e37",
                "bej": "#f5f0d0",
                "vizon": "#c6b299",
                "haki": "#6b705c",
                "camel": "#c19a6b",
                "turuncu": "#ff7b00",
                "tas": "#d6cfc7"
            }

            kod = "#cccccc"
            for key, val in renkler.items():
                if renk_key.endswith(key):
                    kod = val
                    break

            return {"kod": kod, "ad": renk_ad, "key": renk_key}

        # 🔹 Ürün + renk bazında beden toplama
        toplu_liste = {}

        for order in all_orders:
            for l in order.get("lines", []):
                stok = str(l.get("merchantSku") or l.get("productCode") or "BELİRSİZ").strip().upper()
                renk_raw = str(l.get("productColor") or "BELİRSİZ").strip().upper()
                beden = str(l.get("productSize") or "BELİRSİZ").strip().upper()
                urun_adi = STK_TO_NAME.get(stok, str(l.get("productName") or "").strip())

                renk_bilgi = normalize_color_name(renk_raw)

                try:
                    adet = int(l.get("quantity", 1))
                except:
                    adet = 1

                key = (stok, renk_bilgi["key"])

                if key not in toplu_liste:
                    toplu_liste[key] = {
                        "urun_adi": urun_adi,
                        "stok": stok,
                        "renk_ad": renk_bilgi["ad"],
                        "renk_kodu": renk_bilgi["kod"],
                        "adetler": {"S": 0, "M": 0, "L": 0, "XL": 0}
                    }

                toplu_liste[key]["adetler"].setdefault(beden, 0)
                toplu_liste[key]["adetler"][beden] += adet

        SKU_ORDER = ["ETK3I","BMBRTK","HRKA","FDKY","KFTK","BSKLE","SWT3I","ESF3I","ESPE","KMTK","BKTK","KKTK","DYTK","PLR"]

        def sort_key(x):
            try:
                sku_index = SKU_ORDER.index(x["stok"])
            except ValueError:
                sku_index = 999
            return (sku_index, x["renk_ad"].lower())

        tablo = sorted(toplu_liste.values(), key=sort_key)

        return render_template("kargo_toplama.html", tablo=tablo, total=len(tablo))

    except Exception as e:
        import traceback
        print("❌ Kargo Toplama Hatası:", e)
        traceback.print_exc()
        flash(f"Kargo toplama hatası: {e}", "danger")
        return redirect(url_for("dashboard"))

# ---- Excel Raporu İçin Genel Importlar ----
from flask import send_file
import pandas as pd
from io import BytesIO


@app.route("/kargo-raporu")
@login_required
def kargo_raporu():
    import pandas as pd
    from datetime import datetime, timedelta
    import os, tempfile
    from flask import send_file, request
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    now = datetime.now()
    today_str = now.strftime("%d.%m.%Y")
    sheet_name = now.strftime("%Y-%m-%d")

    # ✅ geçici dosya yolu
    tmp_dir = tempfile.gettempdir()
    excel_path = os.path.join(tmp_dir, f"kargo_raporu_{sheet_name}.xlsx")

    # =========================
    # 📅 TARİH FİLTRESİ (URL)
    # =========================
    date_from = request.args.get("from")  # YYYY-MM-DD
    date_to = request.args.get("to")      # YYYY-MM-DD

    q = ShippingLog.query

    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d")
            q = q.filter(ShippingLog.processed_at >= dt_from)
        except:
            pass

    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            q = q.filter(ShippingLog.processed_at < dt_to)
        except:
            pass

    # 🔥 İşleme alınan kayıtlar (DASHBOARD GERÇEĞİ)
    logs = q.order_by(ShippingLog.processed_at.asc()).all()

    rows = []
    for log in logs:
        rows.append({
            "Mağaza": log.supplier_name,
            "Order No": log.order_number,
            "Kargo Barkod": log.tracking_number,
            "Müşteri": log.customer_name,
            "Sipariş Tarihi": log.order_date,
            "İşleme Alınma Tarihi": log.processed_at,
            "Ürün Adı": str(log.product_name),
            "SKU": str(log.sku),
            "Renk": str(log.color),
            "Beden": str(log.size),
            "Adet": int(log.quantity or 1),
        })

    if not rows:
        pd.DataFrame().to_excel(excel_path, index=False)
        return send_file(excel_path, as_attachment=True)

    df = pd.DataFrame(rows)

    # ✅ KESİN DOĞRU SAYI
    total_kargo = len(df)
    teslim_eden = "Baran Özkaya"

    # =========================
    # 📦 EXCEL YAZ
    # =========================
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=2
        )

    # =========================
    # ✍️ ÜST / ALT YAZILAR
    # =========================
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # ÜST
    ws["A1"] = f"Toplam Kargo: {total_kargo}"
    ws["A1"].font = Font(bold=True)

    # ALT
    last_row = ws.max_row + 2

    ws[f"A{last_row}"] = (
        "Yukarıda listelenen kargo paketlerini eksiksiz ve sağlam şekilde teslim aldım."
    )
    ws[f"A{last_row}"].font = Font(bold=True)

    ws[f"A{last_row + 2}"] = f"Teslim Eden: {teslim_eden}"
    ws[f"D{last_row + 2}"] = f"Tarih: {today_str}"
    ws[f"A{last_row + 4}"] = "Teslim Alan: ____________________"

    wb.save(excel_path)

    return send_file(excel_path, as_attachment=True)


# ---- Main ----
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
